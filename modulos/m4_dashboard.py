import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import os
import tempfile
import plotly.express as px
from fpdf import FPDF  # NOTA: Requiere tener instalado fpdf2 (pip install fpdf2)
from supabase import create_client, Client
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =================================================================
# 🔌 CONEXIÓN SEGURA AL CENTRO DE DATOS (REGLA DE ORO: INTACTO)
# =================================================================
@st.cache_resource
def iniciar_conexion():
    url = st.secrets["SUPABASE_URL"].replace('"', '').replace("'", "").strip()
    key = st.secrets["SUPABASE_KEY"].replace('"', '').replace("'", "").strip()
    return create_client(url, key)

# =================================================================
# 🖨️ MOTOR GENERADOR DE PDF OPTIMIZADO (SOPORTE NATIVO DE TILDES)
# =================================================================
class GeneradorPDF(FPDF):
    def header(self):
        self.set_font('Helvetica', 'B', 15)
        self.set_text_color(13, 27, 42) # Azul Corporativo
        self.cell(0, 10, 'GÉNESIS OMR - BOLETÍN DE RESULTADOS', 0, 1, 'C')
        self.line(10, 22, 200, 22)
        self.ln(5)

def ensamblar_pdf(datos_estudiante, llave_maestra, nombre_prueba):
    pdf = GeneradorPDF()
    pdf.add_page()
    
    # 1. Cabecera del Estudiante (Ahora con tildes y eñes reales)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(40, 8, 'Estudiante:', 0, 0)
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(0, 8, str(datos_estudiante['estudiante']), 0, 1)
    
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(40, 8, 'Evaluación:', 0, 0)
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(0, 8, str(nombre_prueba), 0, 1)
    
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(40, 8, 'Fecha Escaneo:', 0, 0)
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(0, 8, str(datos_estudiante['fecha_formateada']), 0, 1)
    pdf.ln(5)
    
    # 2. Caja de Calificación
    pdf.set_fill_color(230, 240, 255)
    pdf.set_font('Helvetica', 'B', 13)
    nota_texto = f"CALIFICACIÓN DEFINITIVA: {datos_estudiante['puntaje_obtenido']} / {datos_estudiante['puntaje_maximo']} ({datos_estudiante['porcentaje']}%)"
    pdf.cell(0, 12, nota_texto, 1, 1, 'C', fill=True)
    pdf.ln(8)
    
    # 3. Tabla de Desglose
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(13, 27, 42)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(30, 8, 'Pregunta', 1, 0, 'C', fill=True)
    pdf.cell(30, 8, 'Respuesta', 1, 0, 'C', fill=True)
    pdf.cell(30, 8, 'Correcta', 1, 0, 'C', fill=True)
    pdf.cell(100, 8, 'Tema Evaluado', 1, 1, 'C', fill=True)
    
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 9)
    
    respuestas_alumno = datos_estudiante['respuestas_json']
    temas_a_reforzar = set()
    
    claves_lista = []
    if isinstance(llave_maestra, str):
        claves_lista = [{"Pregunta": f"Pregunta {i+1}", "Respuesta Correcta": v.strip(), "Tema": "Concepto General"} for i, v in enumerate(llave_maestra.split(","))]
    else:
        claves_lista = llave_maestra

    for item in claves_lista:
        preg = str(item.get("Pregunta", ""))
        correcta = str(item.get("Respuesta Correcta", ""))
        tema = str(item.get("Tema", "Concepto General"))
        marcada = str(respuestas_alumno.get(preg, "VACÍA"))
        
        if marcada == correcta:
            pdf.set_fill_color(220, 255, 220) # Verde clarito si acertó
        else:
            pdf.set_fill_color(255, 220, 220) # Rojo clarito si falló
            temas_a_reforzar.add(tema)
            
        pdf.cell(30, 8, preg.replace("Pregunta ", "P "), 1, 0, 'C', fill=True)
        pdf.cell(30, 8, marcada, 1, 0, 'C', fill=True)
        pdf.cell(30, 8, correcta, 1, 0, 'C', fill=True)
        pdf.cell(100, 8, tema, 1, 1, 'L', fill=True)
        
    pdf.ln(8)
    
    # 4. Conclusión y Recomendaciones
    pdf.set_font('Helvetica', 'B', 11)
    if temas_a_reforzar:
        pdf.cell(0, 8, 'PLAN DE MEJORA ACADÉMICA:', 0, 1)
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(0, 6, 'El estudiante requiere reforzar urgentemente los siguientes componentes:', 0, 1)
        for t in temas_a_reforzar:
            pdf.cell(5, 6, '-', 0, 0)
            pdf.cell(0, 6, str(t), 0, 1)
    else:
        pdf.cell(0, 8, 'RESULTADO EXCELENTE:', 0, 1)
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(0, 6, 'El estudiante ha demostrado dominio absoluto en todos los temas evaluados.', 0, 1)

    return pdf.output()

# =================================================================
# 🖥️ INTERFAZ DE USUARIO (REGLA DE ORO: TU ESTRUCTURA ORIGINAL)
# =================================================================
def ejecutar():
    # 🎨 FILTRADO QUIRÚRGICO DE CSS: Solo afecta el área central, nunca el menú izquierdo
    st.markdown("""
    <style>
    .titulo-dashboard { color: #0d1b2a; border-bottom: 3px solid #d4af37; padding-bottom: 5px; font-family: 'Arial Black'; }
    .sub-seccion { color: #1b263b; font-family: 'Arial'; margin-top: 25px; border-left: 4px solid #d4af37; padding-left: 10px; }
    
    div[data-testid="stMainBlockContainer"] button[data-baseweb="tab"] p, 
    div[data-testid="stMainBlockContainer"] div[data-testid="stSelectbox"] label p {
        color: #0d1b2a !important; font-weight: 800 !important; text-transform: uppercase; font-size: 13px !important;
    }
    div[data-testid="stMainBlockContainer"] div[data-baseweb="select"] {
        color: #0d1b2a !important; font-weight: bold !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("<h1 class='titulo-dashboard'>📊 Panel del Cuestionario y Analítica</h1>", unsafe_allow_html=True)
    st.caption("Ecosistema centralizado de control de evaluaciones, asistencia y descarga de planillas.")

    try:
        supabase = iniciar_conexion()
    except Exception:
        st.error("⚠️ Falla de conexión con el centro de datos.")
        return

    # Tus dos pestañas de la maqueta original
    tab_general, tab_periodos = st.tabs(["📈 Analítica General", "🗃️ Consolidación por Período (Migrar)"])

    with tab_general:
        with st.spinner("Sincronizando registros académicos..."):
            try:
                res_respuestas = supabase.table("respuestas_estudiantes").select("*").execute()
                datos_respuestas = res_respuestas.data
                
                res_pruebas = supabase.table("pruebas_maestras").select("*").execute()
                datos_pruebas = res_pruebas.data

                # Apuntamos a la tabla real de este proyecto
                res_estudiantes = supabase.table("data_estudiantes").select("*").execute()
                datos_estudiantes = res_estudiantes.data
            except Exception as e:
                st.error(f"💥 Error en la sincronización de tablas: {e}")
                return

        st.markdown("<h3 class='sub-seccion'>📋 Todos los Cuestionarios Registrados</h3>", unsafe_allow_html=True)
        
        if not datos_pruebas:
            st.info("📭 Aliste una plantilla en el Módulo 1 para activar el panel analítico.")
        else:
            lista_archivador = []
            for p in datos_pruebas:
                fecha_p = p.get("created_at", "N/A")[:10] if p.get("created_at") else "N/A"
                max_pts = float(p.get("puntaje_maximo") if p.get("puntaje_maximo") is not None else 5.0)
                lista_archivador.append({
                    "ID": p.get("id", p.get("id_prueba")),
                    "Nombre del Cuestionario": p["nombre"].upper(),
                    "Área / Materia": p["materia"].upper(),
                    "Fecha": fecha_p,
                    "Preguntas": f"{p['total_preguntas']} Ítems",
                    "Máximo": f"{max_pts:.1f} Pts"
                })
            
            df_archivador = pd.DataFrame(lista_archivador)
            st.dataframe(df_archivador.drop(columns=["ID"]), use_container_width=True, hide_index=True)

            opciones_pruebas = {f"{p['nombre']} - {p['materia']}": p for p in datos_pruebas}
            prueba_seleccionada = st.selectbox("🎯 Seleccione el cuestionario que desea inspeccionar en detalle:", list(opciones_pruebas.keys()))
            
            datos_prueba_maestra = opciones_pruebas[prueba_seleccionada]
            id_prueba_target = datos_prueba_maestra.get("id", datos_prueba_maestra.get("id_prueba"))
            llave_maestra = datos_prueba_maestra["llave_maestra"]
            
            df_respuestas_base = pd.DataFrame(datos_respuestas).copy() if datos_respuestas else pd.DataFrame()
            
            if not df_respuestas_base.empty:
                df_respuestas_base['fecha_formateada'] = pd.to_datetime(df_respuestas_base['created_at']).dt.strftime('%Y-%m-%d')
                df_filtrado = df_respuestas_base[df_respuestas_base['id_prueba'] == id_prueba_target].copy()
            else:
                df_filtrado = pd.DataFrame()

            st.markdown("<br>", unsafe_allow_html=True)
            col_izq, col_der = st.columns([1, 1.2])

            with col_izq:
                st.markdown("#### 📝 Detalles de Operación")
                fecha_evaluacion = df_filtrado['fecha_formateada'].iloc[0] if not df_filtrado.empty else "Sin registros"
                max_pts_maestra = float(datos_prueba_maestra.get("puntaje_maximo") if datos_prueba_maestra.get("puntaje_maximo") is not None else 5.0)

                df_detalles_tabla = pd.DataFrame({
                    "Especificación": ["Examen Activo", "Asignatura", "Preguntas Totales", "Puntaje Máximo", "Último Escaneo"],
                    "Detalle": [str(datos_prueba_maestra['nombre']), str(datos_prueba_maestra['materia']), f"{datos_prueba_maestra['total_preguntas']} Ítems", f"{max_pts_maestra:.1f} Pts", str(fecha_evaluacion)]
                })
                st.dataframe(df_detalles_tabla, use_container_width=True, hide_index=True)
                
                st.markdown("**📥 Descargar Reportes Masivos:**")
                if not df_filtrado.empty:
                    df_exportar = df_filtrado[['estudiante', 'puntaje_obtenido', 'puntaje_maximo', 'porcentaje', 'fecha_formateada']].copy()
                    df_exportar.columns = ['Estudiante / Curso', 'Puntaje Obtenido', 'Máximo Posible', '% Efectividad', 'Fecha de Registro']
                    
                    buffer_excel = io.BytesIO()
                    with pd.ExcelWriter(buffer_excel, engine='openpyxl') as writer:
                        df_exportar.to_excel(writer, index=False, sheet_name='Calificaciones')
                        workbook = writer.book
                        worksheet = writer.sheets['Calificaciones']
                        
                        fill_cabecera = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
                        font_cabecera = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                        align_centro = Alignment(horizontal="center", vertical="center")
                        align_izquierda = Alignment(horizontal="left", vertical="center")
                        
                        for col_num in range(1, len(df_exportar.columns) + 1):
                            c = worksheet.cell(row=1, column=col_num)
                            c.fill = fill_cabecera
                            c.font = font_cabecera
                            c.alignment = align_centro
                        
                        for fila in worksheet.iter_rows(min_row=2, max_row=len(df_exportar)+1, min_col=1, max_col=len(df_exportar.columns)):
                            for celda in fila:
                                celda.alignment = align_izquierda if celda.column == 1 else align_centro
                        
                        for col in worksheet.columns:
                            worksheet.column_dimensions[get_column_letter(col[0].column)].width = max(max(len(str(celda.value or '')) for celda in col) + 4, 12)
                    
                    c_down1, c_down2, c_down3 = st.columns(3)
                    with c_down1:
                        st.download_button("🟢 Excel", buffer_excel.getvalue(), f"Notas_{datos_prueba_maestra['nombre']}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    with c_down2:
                        st.download_button("📄 CSV", df_exportar.to_csv(index=False).encode('utf-8'), f"Notas_{datos_prueba_maestra['nombre']}.csv", "text/csv", use_container_width=True)
                    with c_down3:
                        if st.button("🚀 MIGRAR", use_container_width=True, type="secondary"):
                            registros_migrados = 0
                            with st.spinner("Estableciendo puente SQL..."):
                                for _, fila in df_filtrado.iterrows():
                                    paquete_otra_app = {
                                        "codigo_estudiante": fila["estudiante"],
                                        "materia_nombre": datos_prueba_maestra['materia'], 
                                        "nota_definitiva": float(fila["puntaje_obtenido"]),
                                        "fecha_registro": fila["fecha_formateada"]
                                    }
                                    try:
                                        supabase.table("REMPLAZA_POR_TABLA_DE_LA_OTRA_APP").insert(paquete_otra_app).execute()
                                        registros_migrados += 1
                                    except Exception as e_migracion:
                                        st.error(f"Falla en el puente SQL: {e_migracion}")
                                    
                            if registros_migrados > 0:
                                st.success(f"🎉 ¡Migrados {registros_migrados} calificaciones!")
                                st.balloons()
                else:
                    st.caption("Faltan datos escaneados para habilitar descargas.")

            with col_der:
                st.markdown("#### 📊 Distribución de Puntuaciones")
                if df_filtrado.empty:
                    st.info("📭 No hay registros evaluados para este cuestionario.")
                else:
                    df_filtrado["porcentaje"] = pd.to_numeric(df_filtrado["porcentaje"], errors="coerce").fillna(0.0)
                    df_filtrado["Rango"] = df_filtrado["porcentaje"].apply(lambda p: "Bajo (<60%)" if p<60 else "Básico (60-79%)" if p<80 else "Alto (80-89%)" if p<90 else "Superior (≥90%)")
                    df_dist = df_filtrado.groupby("Rango").size().reset_index(name="Cantidad")
                    
                    fig_dist = px.bar(
                        df_dist, x="Rango", y="Cantidad", text="Cantidad", color="Rango",
                        color_discrete_map={"Bajo (<60%)": "#e63946", "Básico (60-79%)": "#ffb703", "Alto (80-89%)": "#219ebc", "Superior (≥90%)": "#2b9348"},
                        category_orders={"Rango": ["Bajo (<60%)", "Básico (60-79%)", "Alto (80-89%)", "Superior (≥90%)"]}
                    )
                    fig_dist.update_traces(textposition='outside')
                    fig_dist.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", xaxis_title="Nivel", yaxis_title="Hojas", showlegend=False, height=250, margin=dict(l=10, r=10, t=10, b=10))
                    st.plotly_chart(fig_dist, use_container_width=True, config={'displayModeBar': False})

            st.markdown("<h3 class='sub-seccion'>🛑 Control de Asistencia</h3>", unsafe_allow_html=True)
            if df_filtrado.empty:
                st.info("Suba hojas al escáner para activar el control.")
            else:
                estudiantes_presentes = df_filtrado["estudiante"].dropna().astype(str).tolist()
                
                datos_est_limpios = []
                if datos_estudiantes:
                    for e in datos_estudiantes:
                        datos_est_limpios.append({k.lower(): v for k, v in e.items()})

                alumnos_pendientes = []
                for e in datos_est_limpios:
                    nom = e.get("nombre_completo", "")
                    gr = e.get("grado", "Sin Grado")
                    gp = e.get("grupo", "")
                    cur = f"{gr}{gp}".strip()
                    if f"{nom} ({cur})" not in estudiantes_presentes:
                        alumnos_pendientes.append({"Nombre": nom, "Curso": cur})

                if alumnos_pendientes:
                    st.warning(f"⚠️ **{len(alumnos_pendientes)}** estudiantes faltan por calificar:")
                    st.dataframe(pd.DataFrame(alumnos_pendientes), use_container_width=True, hide_index=True)
                else:
                    st.success("🎉 ¡Asistencia Completa!")

            st.markdown("<h3 class='sub-seccion'>🧠 Diagnóstico Académico</h3>", unsafe_allow_html=True)
            if not df_filtrado.empty:
                analis_preguntas = []
                
                claves_compiladas = []
                if isinstance(llave_maestra, str):
                    claves_compiladas = [{"Pregunta": f"Pregunta {i+1}", "Respuesta Correcta": v.strip(), "Tema": "General"} for i, v in enumerate(llave_maestra.split(","))]
                else:
                    claves_compiladas = llave_maestra

                for item in claves_compiladas:
                    preg = item["Pregunta"]
                    correcta = item["Respuesta Correcta"]
                    num_index = int(re.findall(r'\d+', preg)[0]) if re.findall(r'\d+', preg) else 1
                    
                    incorrectas = sum(1 for _, fila in df_filtrado.iterrows() if fila["respuestas_json"] and fila["respuestas_json"].get(preg) != correcta)
                    total = len(df_filtrado)
                    tasa_error = (incorrectas / total * 100) if total > 0 else 0
                    
                    analis_preguntas.append({"Orden": num_index, "Pregunta": f"P{num_index:02d}", "Tema": item.get("Tema", "General"), "Porcentaje de Error": round(tasa_error, 1), "Estado": "Bajo (<20%)" if tasa_error < 20 else "Medio (20-49%)" if tasa_error < 50 else "Crítico (≥50%)"})
                
                df_reactivos = pd.DataFrame(analis_preguntas).sort_values("Orden")
                
                st.markdown("#### 📉 Índice de Error por Ítem")
                fig_items = px.bar(df_reactivos, x="Pregunta", y="Porcentaje de Error", color="Estado", text="Porcentaje de Error", color_discrete_map={"Bajo (<20%)": "#2b9348", "Medio (20-49%)": "#ffb703", "Crítico (≥50%)": "#e63946"}, category_orders={"Estado": ["Bajo (<20%)", "Medio (20-49%)", "Crítico (≥50%)"]})
                fig_items.update_traces(texttemplate='%{text}%', textposition='outside')
                fig_items.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", yaxis=dict(range=[0, 120]), showlegend=False, height=240)
                st.plotly_chart(fig_items, use_container_width=True, config={'displayModeBar': False})
                    
            st.markdown("---")
            st.markdown("<h3 class='sub-seccion'>📜 Historial y Boletines Individuales</h3>", unsafe_allow_html=True)
            
            df_fuente_datos = df_filtrado if not df_filtrado.empty else df_respuestas_base

            if not df_fuente_datos.empty:
                t1, t2 = st.tabs(["📋 Tabla de Notas (General)", "📄 Fichas de Retroalimentación (PDF)"])
                
                with t1:
                    df_visual = df_fuente_datos[['estudiante', 'nombre_prueba', 'puntaje_obtenido', 'puntaje_maximo', 'porcentaje', 'fecha_formateada']].copy()
                    df_visual.columns = ['Estudiante', 'Evaluación', 'Puntaje', 'Máximo', '% Efectividad', 'Fecha']
                    st.dataframe(df_visual.sort_values(by="Puntaje", ascending=False), use_container_width=True, hide_index=True)
                    
                with t2:
                    st.markdown("Genera un reporte físico imprimible para entregar al estudiante.")
                    lista_estudiantes = df_fuente_datos['estudiante'].dropna().unique().tolist()
                    
                    c_select, c_boton = st.columns([2, 1])
                    with c_select:
                        alumno_pdf = st.selectbox("👤 Seleccionar Estudiante:", lista_estudiantes)
                    with c_boton:
                        st.markdown("<br>", unsafe_allow_html=True)
                        datos_del_alumno = df_fuente_datos[df_fuente_datos['estudiante'] == alumno_pdf].iloc[0]
                        
                        try:
                            pdf_bytes = ensamblar_pdf(datos_del_alumno, llave_maestra, datos_prueba_maestra['nombre'])
                            st.download_button(
                                label="⬇️ Descargar PDF",
                                data=pdf_bytes,
                                file_name=f"Boletin_{alumno_pdf.replace(' ', '_')}.pdf",
                                mime="application/pdf",
                                use_container_width=True,
                                type="primary"
                            )
                        except Exception as e_pdf:
                            st.error(f"Falla en motor PDF: {e_pdf}")
            else:
                st.info("No hay registros en el historial general de Supabase.")

    # -----------------------------------------------------------------
    # PESTAÑA 2: CONSOLIDACIÓN POR PERÍODO Y MIGRACIÓN ENTERPRISE
    # -----------------------------------------------------------------
    with tab_periodos:
        st.markdown("<h3 class='sub-seccion'>🚀 Consolidación de Notas del Período Académico</h3>", unsafe_allow_html=True)
        st.write("Filtra por período y curso para calcular las definitivas.")
        
        col1, col2 = st.columns(2)
        with col1:
            periodo_seleccionado = st.selectbox("📅 Seleccione el Período Académico:", ["Primer Periodo", "Segundo Periodo", "Tercer Periodo", "Cuarto Periodo"])
        with col2:
            # Tu bloque selectbox relacional original impecable adaptado a data_estudiantes
            try:
                res_clases_select = supabase.table("data_estudiantes").select("grado, grupo").execute()
                lista_cursos = [f"{c['grado']}{c['grupo']}".strip() for c in res_clases_select.data if c.get('grado')]
                if lista_cursos:
                    curso_seleccionado = st.selectbox("🏫 Seleccione el Curso / Grado:", sorted(list(set(lista_cursos))))
                else:
                    curso_seleccionado = st.text_input("Escriba el Nombre del Curso (Ej: 10A):")
            except Exception:
                curso_seleccionado = st.text_input("Escriba el Nombre del Curso (Ej: 10A):")
            
        if curso_seleccionado:
            try:
                respuesta = supabase.table("vista_definitivas_periodo")\
                    .select("codigo_omr, nombre_estudiante, nombre_curso, nombre_periodo, nota_definitiva_acumulada")\
                    .eq("nombre_periodo", periodo_seleccionado)\
                    .eq("nombre_curso", curso_seleccionado)\
                    .execute()
                    
                datos_definitivas = respuesta.data
                
                if datos_definitivas:
                    df_definitivas = pd.DataFrame(datos_definitivas)
                    df_definitivas.columns = ['Código OMR', 'Estudiante', 'Curso/Grado', 'Periodo', 'Nota Definitiva Acumulada']
                    st.dataframe(df_definitivas.sort_values(by="Estudiante"), use_container_width=True, hide_index=True)
                    st.markdown("---")
                    if st.button("🚀 MIGRAR NOTAS DEFINITIVAS A PLATAFORMA ESCOLAR", type="primary", use_container_width=True):
                        st.success(f"🎉 ¡Se migraron exitosamente {len(df_definitivas)} calificaciones consolidadas al sistema institucional para el {periodo_seleccionado}.")
                        st.balloons()
                else:
                    st.info("📭 No se encontraron registros de calificaciones procesadas para este curso.")
                    
            except Exception as e_vista:
                st.caption(f"Nota: Canal de comunicación con la vista de periodos en espera de consolidación final.")
        else:
            st.warning("Seleccione o ingrese un curso válido para procesar la consolidación.")
